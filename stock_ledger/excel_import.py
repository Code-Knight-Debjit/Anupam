from dataclasses import dataclass, field
from decimal import Decimal, InvalidOperation
from io import BytesIO
import hashlib
import os
import zipfile

from django.core.files.base import ContentFile
from django.core.files.storage import default_storage
from django.db import transaction
from django.utils import timezone
from django.utils.text import slugify
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.comments import Comment

from products.models import Category, Product, ProductImage, SubCategory
from .models import Branch, BranchStock, OpeningStock
from .services import recompute_branch_stock

FIXED_HEADERS = [
    'Name/SKU', 'Category', 'SubCategory', 'Branch', 'Opening Quantity', 'Cost Price',
    'Low Stock Threshold', 'Visible', 'MRP', 'Image Serial No',
    'Additional Image Serial No 1', 'Additional Image Serial No 2',
]
IMAGE_EXTENSIONS = {'.jpg', '.jpeg', '.png', '.webp', '.gif', '.bmp'}
BULK_IMAGE_DIR = 'products/bulk'
HEADER_FILL = PatternFill(start_color='0E0E11', end_color='0E0E11', fill_type='solid')
HEADER_FONT = Font(color='FFFFFF', bold=True)

TRUE_VALUES = {'true', '1', 'yes', 'y'}
FALSE_VALUES = {'false', '0', 'no', 'n'}


@dataclass
class ImportResult:
    imported: int = 0
    created_products: int = 0
    created_categories: int = 0
    created_subcategories: int = 0
    skipped: int = 0
    images_attached: int = 0      # product records pointed at a main image
    gallery_images_attached: int = 0  # additional (gallery) images attached
    unique_images: int = 0        # distinct files actually written to storage
    errors: list = field(default_factory=list)      # list of (row_number, reason) — bad data, can't proceed
    duplicates: list = field(default_factory=list)   # list of (row_number, reason) — already added previously
    image_warnings: list = field(default_factory=list)  # list of (row_number, reason) — non-fatal: product imported without its image


def build_products_template_workbook() -> BytesIO:
    """Template for bulk-uploading Products: one row per (Product, Branch)
    pair. Name/SKU, Category and Branch are required; SubCategory is optional
    (auto-created under the row's Category if it's new, same as Category
    itself); Opening Quantity + Cost Price together create that product's
    opening balance at that branch (skipped if it already has one there —
    reported as a duplicate, not double-counted); Low Stock Threshold,
    Visible and MRP apply regardless. Image Serial No, Additional Image
    Serial No 1 and Additional Image Serial No 2 are the optional keys for
    the bulk image upload: the matching file in the images .zip is named
    after each (1 → 1.jpg). The first becomes the product's Main image, the
    other two become its Additional (gallery) images. Anything after
    Additional Image Serial No 2 is a Technical Specification: the header is
    the spec name, the cell below is this product's value — shown on the
    public product page. Add as many of these columns as needed;
    re-uploading an existing product with new spec columns fills them in
    (existing keys are kept, new values win)."""
    wb = Workbook()

    ws = wb.active
    ws.title = 'Products'
    spec_example_headers = ['Bore Diameter', 'Outer Diameter']
    ws.append(FIXED_HEADERS + spec_example_headers)
    for cell in ws[1]:
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal='center')
    ws['A1'].comment = Comment('Product identity — type the part code and brand together as free text, e.g. "6205-2RS Timken". If this exact name isn\'t in the catalogue yet, it will be created automatically.', 'Anupam Bearings')
    ws['B1'].comment = Comment('Product category, e.g. Bearings. Created automatically if it doesn\'t exist yet.', 'Anupam Bearings')
    ws['C1'].comment = Comment('Optional. Created automatically under the Category above if it\'s new.', 'Anupam Bearings')
    ws['D1'].comment = Comment('Must exactly match an existing Branch name or code (case-insensitive) — see the Reference sheet.', 'Anupam Bearings')
    ws['E1'].comment = Comment('Opening balance at this branch, whole number of units (Pcs). Leave blank to skip creating an opening balance for this row.', 'Anupam Bearings')
    ws['F1'].comment = Comment('Price per unit for the opening balance. Required if Opening Quantity is given.', 'Anupam Bearings')
    ws['G1'].comment = Comment('Optional. Low-stock alert threshold at this branch.', 'Anupam Bearings')
    ws['H1'].comment = Comment('Optional, defaults to TRUE. Set to FALSE to hide this product from the public site.', 'Anupam Bearings')
    ws['I1'].comment = Comment('Optional. Maximum Retail Price for this product, e.g. 1250.00. Set once per product — the same value applies across every branch row for that product.', 'Anupam Bearings')
    ws['J1'].comment = Comment('Optional. Key for the bulk image upload — this becomes the product\'s Main image. Put a number here and name the matching file after it in the images .zip (e.g. 1 here, 1.jpg in the zip). Supported: .jpg, .jpeg, .png, .webp, .gif, .bmp.', 'Anupam Bearings')
    ws['K1'].comment = Comment('Optional. Same idea as Image Serial No, but for this product\'s first Additional (gallery) image.', 'Anupam Bearings')
    ws['L1'].comment = Comment('Optional. Same idea as Image Serial No, but for this product\'s second Additional (gallery) image.', 'Anupam Bearings')
    ws['M1'].comment = Comment('Anything after Additional Image Serial No 2 is a Technical Specification: the header is the spec name, the cell below is this product\'s value — shown on the public product page. Add as many of these columns as you need.', 'Anupam Bearings')

    ws.append(['6001-ZZ Timken', 'Bearings', 'Ball Bearings', 'Bengaluru', 100, 45.50, 10, 'TRUE', 1250.00, 1, 2, 3, '12mm', '28mm'])
    for cell in ws[2]:
        cell.font = Font(italic=True, color='888888')

    widths = [28, 16, 16, 14, 16, 12, 18, 10, 12, 16, 16, 16, 16, 16]
    for i, width in enumerate(widths, start=1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = width
    ws.freeze_panes = 'A2'

    ref = wb.create_sheet('Reference')
    ref.append(['Valid Branch Names', 'Branch Code', 'Existing Category Names', 'Existing SubCategory Names'])
    for cell in ref[1]:
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
    branches = list(Branch.objects.filter(is_active=True).order_by('name').values_list('name', 'code'))
    category_names = list(Category.objects.order_by('name').values_list('name', flat=True))
    subcategory_names = list(SubCategory.objects.order_by('name').values_list('name', flat=True))
    for i in range(max(len(branches), len(category_names), len(subcategory_names))):
        row = [
            branches[i][0] if i < len(branches) else '',
            branches[i][1] if i < len(branches) else '',
            category_names[i] if i < len(category_names) else '',
            subcategory_names[i] if i < len(subcategory_names) else '',
        ]
        ref.append(row)
    for i, width in enumerate([18, 14, 22, 22], start=1):
        ref.column_dimensions[ref.cell(row=1, column=i).column_letter].width = width

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def _parse_decimal(value, row_number, field_name, errors):
    try:
        return Decimal(str(value)).quantize(Decimal('0.01'))
    except (InvalidOperation, TypeError):
        errors.append((row_number, f'Invalid {field_name}: {value!r}'))
        return None


def _parse_int(value, row_number, field_name, errors):
    try:
        n = int(value)
        if n < 0:
            raise ValueError
        return n
    except (TypeError, ValueError):
        errors.append((row_number, f'Invalid {field_name}: {value!r}'))
        return None


def _parse_bool(value, default=True):
    if value in (None, ''):
        return default
    text = str(value).strip().lower()
    if text in TRUE_VALUES:
        return True
    if text in FALSE_VALUES:
        return False
    return default


def _unique_slug(model, base):
    slug_base = slugify(base) or 'item'
    slug = slug_base
    i = 1
    while model.objects.filter(slug=slug).exists():
        slug = f'{slug_base}-{i}'
        i += 1
    return slug


def _get_or_create_category(name):
    category = Category.objects.filter(name__iexact=name).first()
    if category:
        return category, False
    category = Category.objects.create(name=name, slug=_unique_slug(Category, name))
    return category, True


def _get_or_create_subcategory(name, category):
    subcategory = SubCategory.objects.filter(name__iexact=name, category=category).first()
    if subcategory:
        return subcategory, False
    subcategory = SubCategory.objects.create(name=name, category=category, slug=_unique_slug(SubCategory, name))
    return subcategory, True


def _get_or_create_product(name, category, subcategory, specifications):
    """Technical Specification columns are merged into the product's existing
    specs (new values win, other existing keys are kept) whether the product
    is new or not, so a re-upload can be used to fill in specs for an
    already-catalogued product."""
    product = Product.objects.filter(name__iexact=name).first()
    created = False
    if product is None:
        product = Product.objects.create(
            name=name, slug=_unique_slug(Product, name), category=category, subcategory=subcategory,
        )
        created = True

    if specifications:
        merged = {**(product.specifications or {}), **specifications}
        if merged != (product.specifications or {}):
            product.specifications = merged
            product.save(update_fields=['specifications'])

    return product, created


def _normalize_serial(value):
    """Serials are matched as text, but Excel hands back numbers — a cell
    holding 1 may arrive as int 1 or float 1.0, and both must match "1.jpg"."""
    if value in (None, ''):
        return ''
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _build_header_map(header_row):
    """Locates the fixed columns by header name rather than position, so a
    workbook built from an older template — one with no MRP / Image Serial No
    columns — still imports correctly instead of reading its first two spec
    columns as those fields. Any header that isn't a fixed one is a spec."""
    fixed_lookup = {h.lower(): h for h in FIXED_HEADERS}
    columns = {}
    spec_columns = []
    for index, raw in enumerate(header_row):
        if raw in (None, ''):
            continue
        label = str(raw).strip()
        key = label.lower()
        if key in fixed_lookup and fixed_lookup[key] not in columns:
            columns[fixed_lookup[key]] = index
        else:
            spec_columns.append((label, index))
    return columns, spec_columns


def _store_shared_image(basename, data, cache):
    """Writes each distinct image to storage exactly once and returns its path.

    The point of the bulk image upload is that many products legitimately share
    one photo, and disk on the host is limited — so products reference a single
    stored file instead of each getting its own byte-for-byte copy. The name is
    derived from a hash of the content, which makes it stable across re-imports:
    uploading the same zip again re-uses the file already on disk rather than
    piling up 1.png, 1_a8f2.png, 1_b31c.png and so on.
    """
    digest = hashlib.sha256(data).hexdigest()[:16]
    if digest in cache:
        return cache[digest]
    ext = os.path.splitext(basename)[1].lower() or '.jpg'
    path = f'{BULK_IMAGE_DIR}/{digest}{ext}'
    if not default_storage.exists(path):
        default_storage.save(path, ContentFile(data))
    cache[digest] = path
    return path


def _read_image_archive(uploaded_zip):
    """Reads the optional images .zip into {serial: (filename, bytes)}, keyed
    on each file's name minus its extension. Returns ({}, error) if the upload
    isn't a readable zip — a bad archive shouldn't abort the whole import."""
    if not uploaded_zip:
        return {}, None
    images = {}
    try:
        with zipfile.ZipFile(uploaded_zip) as archive:
            for info in archive.infolist():
                if info.is_dir():
                    continue
                basename = os.path.basename(info.filename)
                if not basename or basename.startswith('.') or '__MACOSX' in info.filename:
                    continue
                stem, ext = os.path.splitext(basename)
                if ext.lower() not in IMAGE_EXTENSIONS:
                    continue
                serial = stem.strip()
                if serial and serial not in images:
                    images[serial] = (basename, archive.read(info))
    except (zipfile.BadZipFile, OSError):
        return {}, 'Could not read the images .zip — upload a valid zip archive. Products were imported without images.'
    return images, None


@transaction.atomic
def import_products_workbook(uploaded_file, user, filename='', image_archive=None):
    """Parses a Products bulk-upload workbook: Name/SKU | Category |
    SubCategory | Branch | Opening Quantity | Cost Price | Low Stock
    Threshold | Visible, one row per (Product, Branch) pair. Category and
    SubCategory are auto-created if new; Branch must already exist (curated
    centrally, never auto-created). A product that already has an opening
    stock entry at that branch has the stock part of the row skipped as a
    duplicate (not double-counted) — Threshold/Visible still apply."""
    wb = load_workbook(uploaded_file, read_only=True, data_only=True)
    ws = wb.active

    header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), ())
    columns, spec_columns = _build_header_map(header_row)

    images, archive_error = _read_image_archive(image_archive)

    errors = []
    duplicates = []
    image_warnings = []
    valid_rows = []  # (product, branch, quantity, price, threshold, is_new_product)
    created_categories = 0
    created_subcategories = 0
    created_products_total = 0
    images_attached = 0
    gallery_images_attached = 0
    imaged_product_ids = set()
    gallery_imaged_product_ids = set()
    stored_image_paths = {}  # content hash -> storage path, so each image is written once
    if archive_error:
        image_warnings.append((0, archive_error))

    rows = list(ws.iter_rows(min_row=2, values_only=True))
    for idx, row in enumerate(rows, start=2):
        if row is None or all(cell in (None, '') for cell in row):
            continue

        def cell(i):
            return row[i] if i is not None and len(row) > i else None

        def col(header):
            return cell(columns.get(header))

        name = str(col('Name/SKU')).strip() if col('Name/SKU') is not None else ''
        category_name = str(col('Category')).strip() if col('Category') is not None else ''
        subcategory_name = str(col('SubCategory')).strip() if col('SubCategory') is not None else ''
        branch_name = str(col('Branch')).strip() if col('Branch') is not None else ''

        if not name or not category_name or not branch_name:
            errors.append((idx, 'Missing Name/SKU, Category, or Branch.'))
            continue

        branch = Branch.objects.filter(name__iexact=branch_name, is_active=True).first()
        if branch is None:
            branch = Branch.objects.filter(code__iexact=branch_name, is_active=True).first()
        if branch is None:
            errors.append((idx, f'No active branch found matching "{branch_name}".'))
            continue

        quantity = None
        price = None
        if col('Opening Quantity') not in (None, ''):
            quantity = _parse_int(col('Opening Quantity'), idx, 'Opening Quantity', errors)
            if quantity is not None and quantity > 0:
                if col('Cost Price') in (None, ''):
                    errors.append((idx, 'Cost Price is required when Opening Quantity is given.'))
                    continue
                price = _parse_decimal(col('Cost Price'), idx, 'Cost Price', errors)
                if price is None:
                    continue
            else:
                quantity = None

        threshold = None
        if col('Low Stock Threshold') not in (None, ''):
            threshold = _parse_int(col('Low Stock Threshold'), idx, 'Low Stock Threshold', errors)
            if threshold is None:
                continue

        is_visible = _parse_bool(col('Visible'), default=True)

        mrp = None
        if col('MRP') not in (None, ''):
            mrp = _parse_decimal(col('MRP'), idx, 'MRP', errors)
            if mrp is None:
                continue

        specifications = {}
        for spec_name, spec_index in spec_columns:
            value = cell(spec_index)
            if value not in (None, ''):
                specifications[spec_name] = str(value).strip()

        category, category_created = _get_or_create_category(category_name)
        if category_created:
            created_categories += 1

        subcategory = None
        if subcategory_name:
            subcategory, subcategory_created = _get_or_create_subcategory(subcategory_name, category)
            if subcategory_created:
                created_subcategories += 1

        product, is_new_product = _get_or_create_product(name, category, subcategory, specifications)
        if is_new_product:
            created_products_total += 1
        if product.is_visible != is_visible:
            product.is_visible = is_visible
            product.save(update_fields=['is_visible'])

        if mrp is not None and product.mrp != mrp:
            product.mrp = mrp
            product.save(update_fields=['mrp'])

        # One product can span several branch rows — only attach its image once.
        serial = _normalize_serial(col('Image Serial No'))
        if serial and product.id not in imaged_product_ids:
            match = images.get(serial)
            if match:
                image_name, image_bytes = match
                # Assign the shared path rather than product.image.save(), which
                # would write a fresh copy of the bytes for every product.
                path = _store_shared_image(image_name, image_bytes, stored_image_paths)
                if product.image.name != path:
                    product.image.name = path
                    product.save(update_fields=['image'])
                imaged_product_ids.add(product.id)
                images_attached += 1
            elif image_archive and not archive_error:
                image_warnings.append((idx, f'No image named "{serial}" found in the zip for "{name}" — imported without an image.'))

        # Same one-shot-per-product rule as the main image, but for the
        # Additional (gallery) slots — a re-upload replaces the product's
        # existing gallery rather than piling more on top of it.
        additional_serials = [
            _normalize_serial(col('Additional Image Serial No 1')),
            _normalize_serial(col('Additional Image Serial No 2')),
        ]
        additional_serials = [s for s in additional_serials if s]
        if additional_serials and product.id not in gallery_imaged_product_ids:
            gallery_imaged_product_ids.add(product.id)
            gallery_paths = []
            for add_serial in additional_serials:
                match = images.get(add_serial)
                if match:
                    image_name, image_bytes = match
                    gallery_paths.append(_store_shared_image(image_name, image_bytes, stored_image_paths))
                elif image_archive and not archive_error:
                    image_warnings.append((idx, f'No image named "{add_serial}" found in the zip for "{name}" — additional image skipped.'))
            if gallery_paths:
                ProductImage.objects.filter(product=product).delete()
                for order, path in enumerate(gallery_paths, start=1):
                    gallery_image = ProductImage(product=product, order=order)
                    gallery_image.image.name = path
                    gallery_image.save()
                    gallery_images_attached += 1

        if quantity and price is not None:
            if OpeningStock.objects.filter(product=product, branch=branch, is_deleted=False).exists():
                duplicates.append((idx, f'"{name}" at {branch.name} already has an opening stock — skipped to avoid double-counting.'))
                if threshold is not None:
                    BranchStock.objects.update_or_create(
                        product=product, branch=branch, defaults={'low_stock_threshold': threshold},
                    )
                continue
            valid_rows.append((product, branch, quantity, price, threshold))
        elif threshold is not None:
            BranchStock.objects.update_or_create(
                product=product, branch=branch, defaults={'low_stock_threshold': threshold},
            )

    result = ImportResult(
        skipped=len(errors) + len(duplicates), errors=errors, duplicates=duplicates,
        created_categories=created_categories, created_subcategories=created_subcategories,
        created_products=created_products_total,
        images_attached=images_attached, gallery_images_attached=gallery_images_attached,
        image_warnings=image_warnings, unique_images=len(stored_image_paths),
    )

    affected_pairs = set()
    for product, branch, quantity, price, threshold in valid_rows:
        OpeningStock.objects.create(
            product=product, branch=branch, quantity=quantity, price=price,
            effective_date=timezone.localdate(), created_by=user,
        )
        affected_pairs.add((product.id, branch.id))
        result.imported += 1

    for product_id, branch_id in affected_pairs:
        recompute_branch_stock(product_id, branch_id)

    if valid_rows:
        for product, branch, quantity, price, threshold in valid_rows:
            if threshold is not None:
                BranchStock.objects.filter(product=product, branch=branch).update(low_stock_threshold=threshold)

    return result
