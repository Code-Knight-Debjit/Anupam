import json
import shutil
import tempfile
import zipfile
from datetime import date, datetime
from decimal import Decimal
from io import BytesIO

from django.contrib.auth.models import User
from django.core.files.uploadedfile import SimpleUploadedFile
from django.db import IntegrityError, transaction
from django.test import TestCase, override_settings
from django.urls import reverse
from django.utils import timezone
from openpyxl import Workbook

from django.core.files.storage import default_storage

from products.models import Category, Product, ProductImage, SubCategory
from .excel_import import BULK_IMAGE_DIR, import_products_workbook
from .models import Branch, BranchStock, OpeningStock, StockPurchase, StockSale, StockTransfer, UserProfile
from .services import historical_stock_matrix, recompute_branch_stock, latest_unit_price


def make_products_workbook_upload(rows, filename='upload.xlsx', extra_headers=None):
    """rows: list of [Name/SKU, Category, SubCategory, Branch, Opening Quantity, Cost Price, Low Stock Threshold, Visible, *specs]."""
    wb = Workbook()
    ws = wb.active
    headers = ['Name/SKU', 'Category', 'SubCategory', 'Branch', 'Opening Quantity', 'Cost Price', 'Low Stock Threshold', 'Visible']
    if extra_headers:
        headers += extra_headers
    ws.append(headers)
    for row in rows:
        ws.append(row)
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return SimpleUploadedFile(filename, buf.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


def make_full_workbook_upload(rows, filename='upload.xlsx', extra_headers=None):
    """Current template shape, including the MRP and Image Serial No columns.
    rows: [Name/SKU, Category, SubCategory, Branch, Opening Quantity, Cost Price,
    Low Stock Threshold, Visible, MRP, Image Serial No, *specs]."""
    wb = Workbook()
    ws = wb.active
    headers = ['Name/SKU', 'Category', 'SubCategory', 'Branch', 'Opening Quantity',
               'Cost Price', 'Low Stock Threshold', 'Visible', 'MRP', 'Image Serial No']
    if extra_headers:
        headers += extra_headers
    ws.append(headers)
    for row in rows:
        ws.append(row)
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return SimpleUploadedFile(filename, buf.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


def make_gallery_workbook_upload(rows, filename='upload.xlsx', extra_headers=None):
    """Same shape as make_full_workbook_upload but also includes the two
    Additional Image Serial No columns.
    rows: [Name/SKU, Category, SubCategory, Branch, Opening Quantity, Cost Price,
    Low Stock Threshold, Visible, MRP, Image Serial No, Additional Image Serial No 1,
    Additional Image Serial No 2, *specs]."""
    wb = Workbook()
    ws = wb.active
    headers = ['Name/SKU', 'Category', 'SubCategory', 'Branch', 'Opening Quantity',
               'Cost Price', 'Low Stock Threshold', 'Visible', 'MRP', 'Image Serial No',
               'Additional Image Serial No 1', 'Additional Image Serial No 2']
    if extra_headers:
        headers += extra_headers
    ws.append(headers)
    for row in rows:
        ws.append(row)
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return SimpleUploadedFile(filename, buf.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


# Smallest valid PNG — enough for ImageField to store without Pillow complaints.
TINY_PNG = bytes.fromhex(
    '89504e470d0a1a0a0000000d4948445200000001000000010806000000'
    '1f15c4890000000a49444154789c6360000002000100ffff0300000600'
    '05572bd8b40000000049454e44ae426082'
)


def make_image_zip(names, filename='images.zip'):
    """names: list of file names to place in the archive, e.g. ['1.jpg', '2.png']."""
    buf = BytesIO()
    with zipfile.ZipFile(buf, 'w') as archive:
        for name in names:
            archive.writestr(name, TINY_PNG)
    buf.seek(0)
    return SimpleUploadedFile(filename, buf.read(), content_type='application/zip')


def make_user(username, role, branch=None, is_staff=True):
    user = User.objects.create_user(username=username, password='testpass123', is_staff=is_staff)
    UserProfile.objects.filter(user=user).delete()
    UserProfile.objects.create(user=user, role=role, branch=branch)
    return user


class StockLedgerTestBase(TestCase):
    def setUp(self):
        # Bengaluru/Chennai are seeded by the stock_ledger data migration.
        self.blr = Branch.objects.get(code='BLR')
        self.maa = Branch.objects.get(code='MAA')
        self.category = Category.objects.create(name='Bearings', slug='bearings')
        self.product = Product.objects.create(
            category=self.category, name='6001-ZZ Timken', slug='6001-zz-timken',
        )
        self.admin = make_user('admin1', UserProfile.ADMIN)
        self.blr_staff = make_user('blrstaff', UserProfile.BRANCH_STAFF, branch=self.blr)
        self.maa_staff = make_user('maastaff', UserProfile.BRANCH_STAFF, branch=self.maa)
        self.viewer = make_user('viewer1', UserProfile.VIEWER)


class BranchScopingTests(StockLedgerTestBase):
    def test_branch_staff_cannot_delete_other_branch_purchase(self):
        entry = StockPurchase.objects.create(
            product=self.product, branch=self.maa, quantity=5, price=20, created_by=self.admin,
        )
        self.client.force_login(self.blr_staff)
        url = reverse('dashboard:stock_ledger:purchase_delete', args=[entry.pk])
        response = self.client.post(url)
        self.assertEqual(response.status_code, 403)
        entry.refresh_from_db()
        self.assertFalse(entry.is_deleted)

    def test_branch_staff_cannot_reach_admin_only_dashboard_sections(self):
        self.client.force_login(self.blr_staff)
        for name in ['dashboard:home', 'dashboard:products', 'dashboard:categories', 'dashboard:subcategories']:
            response = self.client.get(reverse(name))
            self.assertEqual(response.status_code, 403, f'{name} should 403 for Branch Staff')

    def test_branch_staff_cannot_reach_branch_or_user_management(self):
        self.client.force_login(self.blr_staff)
        response = self.client.get(reverse('dashboard:stock_ledger:branches'))
        self.assertEqual(response.status_code, 403)
        response = self.client.get(reverse('dashboard:stock_ledger:users'))
        self.assertEqual(response.status_code, 403)


class ViewerMutationTests(StockLedgerTestBase):
    def test_viewer_cannot_create_transfer(self):
        self.client.force_login(self.viewer)
        url = reverse('dashboard:stock_ledger:transfer_create')
        response = self.client.post(url, {
            'product': self.product.id, 'from_branch': self.blr.id, 'to_branch': self.maa.id, 'quantity': 5,
        })
        self.assertEqual(response.status_code, 403)

    def test_branch_staff_cannot_create_transfer(self):
        self.client.force_login(self.blr_staff)
        url = reverse('dashboard:stock_ledger:transfer_create')
        response = self.client.post(url, {
            'product': self.product.id, 'from_branch': self.blr.id, 'to_branch': self.maa.id, 'quantity': 5,
        })
        self.assertEqual(response.status_code, 403)


class TransferLifecycleTests(StockLedgerTestBase):
    def setUp(self):
        super().setUp()
        OpeningStock.objects.create(
            product=self.product, branch=self.blr, quantity=100, price=10, created_by=self.admin,
        )
        recompute_branch_stock(self.product.id, self.blr.id)

    def test_full_lifecycle_moves_stock_exactly_once(self):
        self.client.force_login(self.admin)
        create_url = reverse('dashboard:stock_ledger:transfer_create')
        self.client.post(create_url, {
            'product': self.product.id, 'from_branch': self.blr.id, 'to_branch': self.maa.id, 'quantity': 30,
        })
        transfer = StockTransfer.objects.get()
        self.assertEqual(transfer.status, StockTransfer.PENDING)

        blr_stock = BranchStock.objects.get(product=self.product, branch=self.blr)
        self.assertEqual(blr_stock.quantity, 100)

        # Dispatch as sending branch staff
        self.client.force_login(self.blr_staff)
        dispatch_url = reverse('dashboard:stock_ledger:transfer_dispatch', args=[transfer.pk])
        response = self.client.post(dispatch_url)
        self.assertEqual(response.json()['success'], True)

        blr_stock.refresh_from_db()
        self.assertEqual(blr_stock.quantity, 70)
        maa_stock, _ = BranchStock.objects.get_or_create(product=self.product, branch=self.maa)
        self.assertEqual(maa_stock.quantity, 0)  # in transit, not yet received

        # Double-dispatch (retry/double-click) must not double-deduct
        response = self.client.post(dispatch_url)
        self.assertEqual(response.json()['success'], False)
        blr_stock.refresh_from_db()
        self.assertEqual(blr_stock.quantity, 70)

        # Receive (fully) as receiving branch staff
        self.client.force_login(self.maa_staff)
        receive_url = reverse('dashboard:stock_ledger:transfer_record_receipt', args=[transfer.pk])
        response = self.client.post(receive_url, data=json.dumps({'quantity': 30}), content_type='application/json')
        body = response.json()
        self.assertTrue(body['success'])
        self.assertEqual(body['status'], StockTransfer.RECEIVED)
        maa_stock.refresh_from_db()
        self.assertEqual(maa_stock.quantity, 30)

        # A further receipt attempt must not double-add (transfer already closed)
        response = self.client.post(receive_url, data=json.dumps({'quantity': 1}), content_type='application/json')
        self.assertEqual(response.json()['success'], False)
        maa_stock.refresh_from_db()
        self.assertEqual(maa_stock.quantity, 30)

    def test_receiving_branch_staff_cannot_dispatch_someone_elses_transfer(self):
        self.client.force_login(self.admin)
        self.client.post(reverse('dashboard:stock_ledger:transfer_create'), {
            'product': self.product.id, 'from_branch': self.blr.id, 'to_branch': self.maa.id, 'quantity': 10,
        })
        transfer = StockTransfer.objects.get()

        self.client.force_login(self.maa_staff)
        response = self.client.post(reverse('dashboard:stock_ledger:transfer_dispatch', args=[transfer.pk]))
        self.assertEqual(response.status_code, 403)


class RecomputeBranchStockTests(StockLedgerTestBase):
    def test_recompute_reflects_edit_and_soft_delete(self):
        entry = StockPurchase.objects.create(
            product=self.product, branch=self.blr, quantity=20, price=5, created_by=self.admin,
        )
        recompute_branch_stock(self.product.id, self.blr.id)
        stock = BranchStock.objects.get(product=self.product, branch=self.blr)
        self.assertEqual(stock.quantity, 20)

        entry.quantity = 50
        entry.save()
        recompute_branch_stock(self.product.id, self.blr.id)
        stock.refresh_from_db()
        self.assertEqual(stock.quantity, 50)

        entry.soft_delete(self.admin)
        recompute_branch_stock(self.product.id, self.blr.id)
        stock.refresh_from_db()
        self.assertEqual(stock.quantity, 0)

    def test_sale_reduces_computed_stock(self):
        StockPurchase.objects.create(product=self.product, branch=self.blr, quantity=20, price=5, created_by=self.admin)
        recompute_branch_stock(self.product.id, self.blr.id)
        StockSale.objects.create(
            product=self.product, branch=self.blr, quantity=6, price=5, selling_price=8, created_by=self.admin,
        )
        recompute_branch_stock(self.product.id, self.blr.id)
        stock = BranchStock.objects.get(product=self.product, branch=self.blr)
        self.assertEqual(stock.quantity, 14)


class ProductNameIdentityTests(StockLedgerTestBase):
    """Brand/SKU are gone — Product.name is now the sole, unique identity field."""

    def test_duplicate_name_rejected(self):
        with self.assertRaises(IntegrityError):
            with transaction.atomic():
                Product.objects.create(
                    category=self.category, name=self.product.name, slug='dup-6001-zz',
                )

    def test_different_name_allowed(self):
        Product.objects.create(category=self.category, name='6205-2RS NSK', slug='6205-2rs-nsk')
        self.assertEqual(Product.objects.count(), 2)


class SaleStockLimitTests(StockLedgerTestBase):
    def setUp(self):
        super().setUp()
        OpeningStock.objects.create(product=self.product, branch=self.blr, quantity=10, price=5, created_by=self.admin)
        recompute_branch_stock(self.product.id, self.blr.id)

    def test_sale_exceeding_stock_is_rejected(self):
        self.client.force_login(self.blr_staff)
        response = self.client.post(reverse('dashboard:stock_ledger:sale_add'), {
            'sale_date': '2026-01-01', 'customer': '', 'notes': '',
            'item_product': [str(self.product.id)], 'item_quantity': ['11'],
            'item_selling_price': ['8'],
        })
        self.assertEqual(response.status_code, 200)  # re-rendered form, not redirected
        self.assertEqual(StockSale.objects.count(), 0)
        stock = BranchStock.objects.get(product=self.product, branch=self.blr)
        self.assertEqual(stock.quantity, 10)

    def test_sale_within_stock_succeeds_and_never_goes_negative(self):
        self.client.force_login(self.blr_staff)
        response = self.client.post(reverse('dashboard:stock_ledger:sale_add'), {
            'sale_date': '2026-01-01', 'customer': '', 'notes': '',
            'item_product': [str(self.product.id)], 'item_quantity': ['10'],
            'item_selling_price': ['8'],
        })
        self.assertEqual(response.status_code, 302)
        stock = BranchStock.objects.get(product=self.product, branch=self.blr)
        self.assertEqual(stock.quantity, 0)
        self.assertGreaterEqual(stock.quantity, 0)

    def test_transfer_exceeding_stock_is_rejected(self):
        self.client.force_login(self.admin)
        response = self.client.post(reverse('dashboard:stock_ledger:transfer_create'), {
            'product': self.product.id, 'from_branch': self.blr.id, 'to_branch': self.maa.id, 'quantity': 999,
        })
        self.assertEqual(response.status_code, 200)
        self.assertEqual(StockTransfer.objects.count(), 0)


class TransferPartialReceiptTests(StockLedgerTestBase):
    def setUp(self):
        super().setUp()
        OpeningStock.objects.create(product=self.product, branch=self.blr, quantity=100, price=10, created_by=self.admin)
        recompute_branch_stock(self.product.id, self.blr.id)
        self.client.force_login(self.admin)
        self.client.post(reverse('dashboard:stock_ledger:transfer_create'), {
            'product': self.product.id, 'from_branch': self.blr.id, 'to_branch': self.maa.id, 'quantity': 50,
        })
        self.transfer = StockTransfer.objects.get()
        self.client.force_login(self.blr_staff)
        self.client.post(reverse('dashboard:stock_ledger:transfer_dispatch', args=[self.transfer.pk]))
        self.transfer.refresh_from_db()

    def _receive(self, quantity):
        self.client.force_login(self.maa_staff)
        return self.client.post(
            reverse('dashboard:stock_ledger:transfer_record_receipt', args=[self.transfer.pk]),
            data=json.dumps({'quantity': quantity}), content_type='application/json',
        )

    def test_partial_then_full_receipt_closes_transfer(self):
        response = self._receive(20)
        body = response.json()
        self.assertTrue(body['success'])
        self.assertEqual(body['status'], StockTransfer.PARTIALLY_RECEIVED)
        self.assertEqual(body['remaining_quantity'], 30)
        maa_stock = BranchStock.objects.get(product=self.product, branch=self.maa)
        self.assertEqual(maa_stock.quantity, 20)

        response = self._receive(30)
        body = response.json()
        self.assertTrue(body['success'])
        self.assertEqual(body['status'], StockTransfer.RECEIVED)
        maa_stock.refresh_from_db()
        self.assertEqual(maa_stock.quantity, 50)

    def test_receipt_cannot_exceed_remaining(self):
        response = self._receive(999)
        self.assertFalse(response.json()['success'])
        self.assertEqual(StockTransfer.objects.get(pk=self.transfer.pk).received_quantity, 0)

    def test_not_received_moves_to_issue_reported_and_admin_resolves(self):
        self.client.force_login(self.maa_staff)
        response = self.client.post(
            reverse('dashboard:stock_ledger:transfer_report_not_received', args=[self.transfer.pk])
        )
        self.assertTrue(response.json()['success'])
        self.transfer.refresh_from_db()
        self.assertEqual(self.transfer.status, StockTransfer.ISSUE_REPORTED)

        # Branch staff cannot resolve issues — Admin only.
        response = self.client.post(
            reverse('dashboard:stock_ledger:transfer_resolve_issue', args=[self.transfer.pk])
        )
        self.assertEqual(response.status_code, 403)

        self.client.force_login(self.admin)
        response = self.client.post(
            reverse('dashboard:stock_ledger:transfer_resolve_issue', args=[self.transfer.pk])
        )
        body = response.json()
        self.assertTrue(body['success'])
        self.assertEqual(body['status'], StockTransfer.DISPATCHED)


class ProductQuickCreateTests(StockLedgerTestBase):
    def test_quick_create_defaults_to_uncategorized(self):
        self.client.force_login(self.blr_staff)
        response = self.client.post(reverse('dashboard:stock_ledger:product_quick_create'), {
            'name': '6205-2RS Timken',
        })
        body = response.json()
        self.assertTrue(body['success'])
        product = Product.objects.get(name='6205-2RS Timken')
        self.assertEqual(product.category.slug, 'uncategorized')

    def test_quick_create_is_idempotent_for_existing_name(self):
        self.client.force_login(self.blr_staff)
        first = self.client.post(reverse('dashboard:stock_ledger:product_quick_create'), {
            'name': self.product.name,
        }).json()
        self.assertEqual(first['product']['id'], self.product.id)
        self.assertEqual(Product.objects.filter(name=self.product.name).count(), 1)

    def test_viewer_cannot_quick_create(self):
        self.client.force_login(self.viewer)
        response = self.client.post(reverse('dashboard:stock_ledger:product_quick_create'), {'name': 'X'})
        self.assertEqual(response.status_code, 403)


class ProductSearchTests(StockLedgerTestBase):
    def test_search_restricted_to_in_stock_branch(self):
        OpeningStock.objects.create(product=self.product, branch=self.blr, quantity=5, price=1, created_by=self.admin)
        recompute_branch_stock(self.product.id, self.blr.id)

        self.client.force_login(self.blr_staff)
        url = reverse('dashboard:stock_ledger:product_search')
        response = self.client.get(url, {'q': '6001', 'in_stock_branch': self.blr.id})
        self.assertEqual(len(response.json()['results']), 1)

        response = self.client.get(url, {'q': '6001', 'in_stock_branch': self.maa.id})
        self.assertEqual(len(response.json()['results']), 0)


class ProductSearchAvailableQtyTests(StockLedgerTestBase):
    """Item 12 — the picker shows how much is on hand before staff commit to a qty."""

    def test_available_qty_returned_for_stock_restricted_search(self):
        OpeningStock.objects.create(product=self.product, branch=self.blr, quantity=42, price=1, created_by=self.admin)
        recompute_branch_stock(self.product.id, self.blr.id)

        self.client.force_login(self.blr_staff)
        response = self.client.get(
            reverse('dashboard:stock_ledger:product_search'),
            {'q': '6001', 'in_stock_branch': self.blr.id},
        )
        results = response.json()['results']
        self.assertEqual(len(results), 1)
        self.assertEqual(results[0]['available_qty'], 42)

    def test_available_qty_is_null_when_search_is_not_branch_restricted(self):
        # Purchases search unrestricted — there's no single branch to report.
        self.client.force_login(self.admin)
        response = self.client.get(reverse('dashboard:stock_ledger:product_search'), {'q': '6001'})
        results = response.json()['results']
        self.assertEqual(len(results), 1)
        self.assertIsNone(results[0]['available_qty'])


@override_settings(MEDIA_ROOT=tempfile.mkdtemp())
class ProductImportMrpAndImageTests(StockLedgerTestBase):
    """Items 9 and 10 — the MRP column and zip-based image matching."""

    @classmethod
    def tearDownClass(cls):
        shutil.rmtree(cls._overridden_settings['MEDIA_ROOT'], ignore_errors=True)
        super().tearDownClass()

    def test_mrp_column_is_imported(self):
        upload = make_full_workbook_upload([
            ['MRP-SKU-1 Timken', self.category.name, '', self.blr.name, 10, 5, '', '', 1250.50, ''],
        ])
        import_products_workbook(upload, self.admin)
        product = Product.objects.get(name='MRP-SKU-1 Timken')
        self.assertEqual(product.mrp, Decimal('1250.50'))

    def test_blank_mrp_leaves_field_null(self):
        upload = make_full_workbook_upload([
            ['MRP-SKU-2 Timken', self.category.name, '', self.blr.name, 10, 5, '', '', '', ''],
        ])
        import_products_workbook(upload, self.admin)
        self.assertIsNone(Product.objects.get(name='MRP-SKU-2 Timken').mrp)

    def test_image_is_attached_by_serial_number(self):
        upload = make_full_workbook_upload([
            ['IMG-SKU-1 Timken', self.category.name, '', self.blr.name, 10, 5, '', '', '', 1],
        ])
        result = import_products_workbook(upload, self.admin, image_archive=make_image_zip(['1.png']))

        self.assertEqual(result.images_attached, 1)
        self.assertEqual(result.image_warnings, [])
        self.assertTrue(Product.objects.get(name='IMG-SKU-1 Timken').image)

    def test_serial_with_no_matching_file_warns_but_still_imports_product(self):
        upload = make_full_workbook_upload([
            ['IMG-SKU-2 Timken', self.category.name, '', self.blr.name, 10, 5, '', '', '', 99],
        ])
        result = import_products_workbook(upload, self.admin, image_archive=make_image_zip(['1.png']))

        self.assertEqual(result.images_attached, 0)
        self.assertEqual(len(result.image_warnings), 1)
        self.assertEqual(len(result.errors), 0)
        product = Product.objects.get(name='IMG-SKU-2 Timken')
        self.assertFalse(product.image)

    def test_import_without_zip_produces_no_image_warnings(self):
        upload = make_full_workbook_upload([
            ['IMG-SKU-3 Timken', self.category.name, '', self.blr.name, 10, 5, '', '', '', 1],
        ])
        result = import_products_workbook(upload, self.admin)
        self.assertEqual(result.images_attached, 0)
        self.assertEqual(result.image_warnings, [])

    def test_image_attached_once_when_product_spans_several_branch_rows(self):
        upload = make_full_workbook_upload([
            ['IMG-SKU-4 Timken', self.category.name, '', self.blr.name, 10, 5, '', '', '', 1],
            ['IMG-SKU-4 Timken', self.category.name, '', self.maa.name, 4, 5, '', '', '', 1],
        ])
        result = import_products_workbook(upload, self.admin, image_archive=make_image_zip(['1.png']))
        self.assertEqual(result.images_attached, 1)

    def test_products_sharing_a_serial_reference_one_stored_file(self):
        """The whole point of the bulk image upload: reusing one photo across
        many products must cost one file on disk, not one copy per product."""
        upload = make_full_workbook_upload([
            ['SHARE-1 Timken', self.category.name, '', self.blr.name, 1, 5, '', '', '', 7],
            ['SHARE-2 Timken', self.category.name, '', self.blr.name, 1, 5, '', '', '', 7],
            ['SHARE-3 Timken', self.category.name, '', self.blr.name, 1, 5, '', '', '', 7],
        ])
        result = import_products_workbook(upload, self.admin, image_archive=make_image_zip(['7.png']))

        self.assertEqual(result.images_attached, 3)
        self.assertEqual(result.unique_images, 1)

        names = {Product.objects.get(name=f'SHARE-{i} Timken').image.name for i in (1, 2, 3)}
        self.assertEqual(len(names), 1, 'all three products must point at the same stored file')

        path = names.pop()
        self.assertTrue(path)
        stored = [f for f in default_storage.listdir(BULK_IMAGE_DIR)[1]]
        self.assertEqual(len(stored), 1, f'expected exactly one file on disk, found {stored}')

    def test_reimporting_the_same_zip_does_not_duplicate_files_on_disk(self):
        rows = [['REIMP-1 Timken', self.category.name, '', self.blr.name, '', '', '', '', '', 3]]
        import_products_workbook(make_full_workbook_upload(rows), self.admin,
                                 image_archive=make_image_zip(['3.png']))
        first = Product.objects.get(name='REIMP-1 Timken').image.name

        import_products_workbook(make_full_workbook_upload(rows), self.admin,
                                 image_archive=make_image_zip(['3.png']))
        second = Product.objects.get(name='REIMP-1 Timken').image.name

        self.assertEqual(first, second)
        stored = default_storage.listdir(BULK_IMAGE_DIR)[1]
        self.assertEqual(len(stored), 1, f're-import should re-use the stored file, found {stored}')

    def test_unreadable_zip_warns_but_products_still_import(self):
        upload = make_full_workbook_upload([
            ['IMG-SKU-5 Timken', self.category.name, '', self.blr.name, 10, 5, '', '', '', 1],
        ])
        not_a_zip = SimpleUploadedFile('images.zip', b'this is not a zip', content_type='application/zip')
        result = import_products_workbook(upload, self.admin, image_archive=not_a_zip)

        self.assertEqual(result.images_attached, 0)
        self.assertEqual(len(result.image_warnings), 1)
        self.assertTrue(Product.objects.filter(name='IMG-SKU-5 Timken').exists())


@override_settings(MEDIA_ROOT=tempfile.mkdtemp())
class ProductImportAdditionalImagesTests(StockLedgerTestBase):
    """Main image (Image Serial No) plus two Additional (gallery) image slots."""

    @classmethod
    def tearDownClass(cls):
        shutil.rmtree(cls._overridden_settings['MEDIA_ROOT'], ignore_errors=True)
        super().tearDownClass()

    def test_additional_images_attached_as_gallery(self):
        upload = make_gallery_workbook_upload([
            ['GAL-SKU-1 Timken', self.category.name, '', self.blr.name, 10, 5, '', '', '', 1, 2, 3],
        ])
        result = import_products_workbook(
            upload, self.admin, image_archive=make_image_zip(['1.png', '2.png', '3.png']),
        )

        self.assertEqual(result.images_attached, 1)
        self.assertEqual(result.gallery_images_attached, 2)
        self.assertEqual(result.image_warnings, [])
        product = Product.objects.get(name='GAL-SKU-1 Timken')
        self.assertTrue(product.image)
        self.assertEqual(ProductImage.objects.filter(product=product).count(), 2)

    def test_missing_additional_file_warns_but_still_imports(self):
        upload = make_gallery_workbook_upload([
            ['GAL-SKU-2 Timken', self.category.name, '', self.blr.name, 10, 5, '', '', '', 1, 99, ''],
        ])
        result = import_products_workbook(
            upload, self.admin, image_archive=make_image_zip(['1.png']),
        )

        self.assertEqual(result.gallery_images_attached, 0)
        self.assertEqual(len(result.image_warnings), 1)
        product = Product.objects.get(name='GAL-SKU-2 Timken')
        self.assertEqual(ProductImage.objects.filter(product=product).count(), 0)

    def test_gallery_attached_once_when_product_spans_several_branch_rows(self):
        upload = make_gallery_workbook_upload([
            ['GAL-SKU-3 Timken', self.category.name, '', self.blr.name, 10, 5, '', '', '', '', 2, 3],
            ['GAL-SKU-3 Timken', self.category.name, '', self.maa.name, 4, 5, '', '', '', '', 2, 3],
        ])
        result = import_products_workbook(
            upload, self.admin, image_archive=make_image_zip(['2.png', '3.png']),
        )
        self.assertEqual(result.gallery_images_attached, 2)
        product = Product.objects.get(name='GAL-SKU-3 Timken')
        self.assertEqual(ProductImage.objects.filter(product=product).count(), 2)

    def test_reimport_replaces_existing_gallery(self):
        rows_first = [['GAL-SKU-4 Timken', self.category.name, '', self.blr.name, '', '', '', '', '', '', 2, 3]]
        import_products_workbook(
            make_gallery_workbook_upload(rows_first), self.admin,
            image_archive=make_image_zip(['2.png', '3.png']),
        )
        product = Product.objects.get(name='GAL-SKU-4 Timken')
        self.assertEqual(ProductImage.objects.filter(product=product).count(), 2)

        rows_second = [['GAL-SKU-4 Timken', self.category.name, '', self.blr.name, '', '', '', '', '', '', 5, '']]
        import_products_workbook(
            make_gallery_workbook_upload(rows_second), self.admin,
            image_archive=make_image_zip(['5.png']),
        )
        self.assertEqual(ProductImage.objects.filter(product=product).count(), 1)


class ProductImportOldTemplateCompatTests(StockLedgerTestBase):
    """Columns are matched by header name, so a workbook saved from the template
    that predates MRP / Image Serial No must not have its first two spec columns
    misread as those fields."""

    def test_old_template_spec_columns_are_still_read_as_specs(self):
        upload = make_products_workbook_upload(
            [['OLD-TPL-1 Timken', self.category.name, '', self.blr.name, 10, 5, '', '', '12mm', '28mm']],
            extra_headers=['Bore Diameter', 'Outer Diameter'],
        )
        import_products_workbook(upload, self.admin)

        product = Product.objects.get(name='OLD-TPL-1 Timken')
        self.assertEqual(product.specifications.get('Bore Diameter'), '12mm')
        self.assertEqual(product.specifications.get('Outer Diameter'), '28mm')
        self.assertIsNone(product.mrp)

    def test_columns_in_a_different_order_still_import(self):
        wb = Workbook()
        ws = wb.active
        ws.append(['Branch', 'Name/SKU', 'MRP', 'Category', 'Opening Quantity', 'Cost Price'])
        ws.append([self.blr.name, 'REORDER-1 Timken', 999.00, self.category.name, 3, 2])
        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)
        upload = SimpleUploadedFile('reordered.xlsx', buf.read())

        result = import_products_workbook(upload, self.admin)

        self.assertEqual(result.imported, 1)
        self.assertEqual(Product.objects.get(name='REORDER-1 Timken').mrp, Decimal('999.00'))


class ProductBulkImportTests(StockLedgerTestBase):
    def test_new_product_is_auto_created_with_opening_stock(self):
        upload = make_products_workbook_upload([
            ['NEW-SKU-1 Timken', self.category.name, '', self.blr.name, 10, 5, '', ''],
        ])
        result = import_products_workbook(upload, self.admin)

        self.assertEqual(result.imported, 1)
        self.assertEqual(result.created_products, 1)
        self.assertEqual(result.skipped, 0)
        product = Product.objects.get(name='NEW-SKU-1 Timken')
        self.assertEqual(product.category_id, self.category.id)
        self.assertTrue(product.is_visible)
        stock = BranchStock.objects.get(product=product, branch=self.blr)
        self.assertEqual(stock.quantity, 10)

    def test_new_category_and_subcategory_are_auto_created(self):
        upload = make_products_workbook_upload([
            ['NEW-SKU-2 Timken', 'Brand New Category', 'Brand New Sub', self.blr.name, 10, 5, '', ''],
        ])
        result = import_products_workbook(upload, self.admin)

        self.assertEqual(result.created_categories, 1)
        self.assertEqual(result.created_subcategories, 1)
        product = Product.objects.get(name='NEW-SKU-2 Timken')
        self.assertEqual(product.category.name, 'Brand New Category')
        self.assertEqual(product.subcategory.name, 'Brand New Sub')

    def test_threshold_and_visible_apply_without_opening_stock(self):
        upload = make_products_workbook_upload([
            ['NEW-SKU-3 Timken', self.category.name, '', self.blr.name, '', '', 7, 'FALSE'],
        ])
        result = import_products_workbook(upload, self.admin)

        self.assertEqual(result.imported, 0)
        product = Product.objects.get(name='NEW-SKU-3 Timken')
        self.assertFalse(product.is_visible)
        stock = BranchStock.objects.get(product=product, branch=self.blr)
        self.assertEqual(stock.low_stock_threshold, 7)
        self.assertEqual(stock.quantity, 0)

    def test_reupload_of_already_added_product_is_flagged_as_duplicate_not_error(self):
        OpeningStock.objects.create(
            product=self.product, branch=self.blr, quantity=50, price=10, created_by=self.admin,
        )
        recompute_branch_stock(self.product.id, self.blr.id)

        upload = make_products_workbook_upload([
            [self.product.name, self.category.name, '', self.blr.name, 999, 1, '', ''],
        ])
        result = import_products_workbook(upload, self.admin)

        self.assertEqual(result.imported, 0)
        self.assertEqual(len(result.duplicates), 1)
        self.assertEqual(len(result.errors), 0)

        stock = BranchStock.objects.get(product=self.product, branch=self.blr)
        self.assertEqual(stock.quantity, 50)

    def test_unknown_branch_is_skipped_as_error_not_auto_created(self):
        upload = make_products_workbook_upload([
            ['SOME-SKU', self.category.name, '', 'TotallyMadeUpBranch', 10, 5, '', ''],
        ])
        result = import_products_workbook(upload, self.admin)

        self.assertEqual(result.imported, 0)
        self.assertEqual(len(result.errors), 1)
        self.assertFalse(Branch.objects.filter(name='TotallyMadeUpBranch').exists())

    def test_dynamic_spec_columns_are_merged_into_product_specifications(self):
        upload = make_products_workbook_upload(
            [['NEW-SKU-4 Timken', self.category.name, '', self.blr.name, 10, 5, '', '', '12mm', '28mm']],
            extra_headers=['Bore Diameter', 'Outer Diameter'],
        )
        import_products_workbook(upload, self.admin)
        product = Product.objects.get(name='NEW-SKU-4 Timken')
        self.assertEqual(product.specifications.get('Bore Diameter'), '12mm')
        self.assertEqual(product.specifications.get('Outer Diameter'), '28mm')

    def test_spec_columns_fill_in_on_reupload_of_existing_product(self):
        self.assertEqual(self.product.specifications, {})
        upload = make_products_workbook_upload(
            [[self.product.name, self.category.name, '', self.maa.name, '', '', '', '', '15mm']],
            extra_headers=['Bore Diameter'],
        )
        import_products_workbook(upload, self.admin)
        self.product.refresh_from_db()
        self.assertEqual(self.product.specifications.get('Bore Diameter'), '15mm')

    def test_existing_product_without_opening_stock_at_branch_still_imports(self):
        upload = make_products_workbook_upload([
            [self.product.name, self.category.name, '', self.maa.name, 7, 3, '', ''],
        ])
        result = import_products_workbook(upload, self.admin)

        self.assertEqual(result.imported, 1)
        self.assertEqual(result.created_products, 0)  # product already existed
        stock = BranchStock.objects.get(product=self.product, branch=self.maa)
        self.assertEqual(stock.quantity, 7)


class MultiItemPurchaseTests(StockLedgerTestBase):
    def test_single_purchase_creates_one_row_per_item_sharing_supplier_and_notes(self):
        other_product = Product.objects.create(
            category=self.category, name='6205-2RS Timken', slug='6205-2rs-timken',
        )
        self.client.force_login(self.blr_staff)
        response = self.client.post(reverse('dashboard:stock_ledger:purchase_add'), {
            'purchase_date': '2026-02-01', 'supplier': 'Blinkit', 'notes': 'Bulk restock order',
            'item_product': [str(self.product.id), str(other_product.id)],
            'item_quantity': ['20', '15'],
            'item_price': ['10.00', '25.50'],
        })
        self.assertEqual(response.status_code, 302)
        entries = StockPurchase.objects.filter(supplier='Blinkit').order_by('id')
        self.assertEqual(entries.count(), 2)
        self.assertTrue(all(e.notes == 'Bulk restock order' for e in entries))
        self.assertTrue(all(e.purchase_date.isoformat() == '2026-02-01' for e in entries))

        stock1 = BranchStock.objects.get(product=self.product, branch=self.blr)
        stock2 = BranchStock.objects.get(product=other_product, branch=self.blr)
        self.assertEqual(stock1.quantity, 20)
        self.assertEqual(stock2.quantity, 15)

    def test_incomplete_rows_are_silently_skipped(self):
        self.client.force_login(self.blr_staff)
        response = self.client.post(reverse('dashboard:stock_ledger:purchase_add'), {
            'purchase_date': '2026-02-01', 'supplier': 'Blinkit', 'notes': '',
            'item_product': [str(self.product.id), ''],
            'item_quantity': ['5', ''],
            'item_price': ['10.00', ''],
        })
        self.assertEqual(response.status_code, 302)
        self.assertEqual(StockPurchase.objects.filter(supplier='Blinkit').count(), 1)

    def test_no_items_at_all_reports_error_without_crashing(self):
        self.client.force_login(self.blr_staff)
        response = self.client.post(reverse('dashboard:stock_ledger:purchase_add'), {
            'purchase_date': '2026-02-01', 'supplier': 'Blinkit', 'notes': '',
            'item_product': [''], 'item_quantity': [''], 'item_price': [''],
        })
        self.assertEqual(response.status_code, 200)
        self.assertEqual(StockPurchase.objects.count(), 0)


class SaleAddPageTests(StockLedgerTestBase):
    """Regression coverage for the GET-request crash: the template referenced
    `posted` as a filter *argument* (`default:posted.x`), which — unlike a
    plain top-level `{{ posted.x }}` — isn't protected by Django's silent
    missing-variable handling, so it 500'd whenever `posted` wasn't in the
    render context (i.e. every normal page load, not just the error path)."""
    def test_add_sale_page_loads_without_posted_in_context(self):
        self.client.force_login(self.blr_staff)
        response = self.client.get(reverse('dashboard:stock_ledger:sale_add'))
        self.assertEqual(response.status_code, 200)

    def test_add_purchase_page_loads_without_posted_in_context(self):
        self.client.force_login(self.blr_staff)
        response = self.client.get(reverse('dashboard:stock_ledger:purchase_add'))
        self.assertEqual(response.status_code, 200)


class MultiItemSaleTests(StockLedgerTestBase):
    def setUp(self):
        super().setUp()
        self.other_product = Product.objects.create(
            category=self.category, name='6205-2RS Timken', slug='6205-2rs-timken',
        )
        OpeningStock.objects.create(product=self.product, branch=self.blr, quantity=20, price=5, created_by=self.admin)
        OpeningStock.objects.create(product=self.other_product, branch=self.blr, quantity=15, price=8, created_by=self.admin)
        recompute_branch_stock(self.product.id, self.blr.id)
        recompute_branch_stock(self.other_product.id, self.blr.id)

    def test_single_sale_creates_one_row_per_item_sharing_customer_and_notes(self):
        self.client.force_login(self.blr_staff)
        response = self.client.post(reverse('dashboard:stock_ledger:sale_add'), {
            'sale_date': '2026-03-01', 'customer': 'Walk-in', 'notes': 'Counter sale',
            'item_product': [str(self.product.id), str(self.other_product.id)],
            'item_quantity': ['5', '3'],
            'item_selling_price': ['9.00', '14.00'],
        })
        self.assertEqual(response.status_code, 302)
        entries = StockSale.objects.filter(customer='Walk-in').order_by('id')
        self.assertEqual(entries.count(), 2)
        self.assertTrue(all(e.notes == 'Counter sale' for e in entries))

        stock1 = BranchStock.objects.get(product=self.product, branch=self.blr)
        stock2 = BranchStock.objects.get(product=self.other_product, branch=self.blr)
        self.assertEqual(stock1.quantity, 15)  # 20 - 5
        self.assertEqual(stock2.quantity, 12)  # 15 - 3

    def test_cost_price_is_pulled_from_latest_unit_price_not_the_form(self):
        """The Sale form no longer submits a cost price at all — price is
        computed server-side from latest_unit_price(), matching the price the
        product's most recent OpeningStock/Purchase entry recorded."""
        self.client.force_login(self.blr_staff)
        self.client.post(reverse('dashboard:stock_ledger:sale_add'), {
            'sale_date': '2026-03-01', 'customer': 'Walk-in', 'notes': '',
            'item_product': [str(self.product.id)],
            'item_quantity': ['5'],
            'item_selling_price': ['9.00'],
        })
        entry = StockSale.objects.get(customer='Walk-in')
        self.assertEqual(entry.price, latest_unit_price(self.product.id, self.blr.id))
        self.assertEqual(entry.price, 5)

    def test_sum_of_rows_for_same_product_cannot_exceed_stock(self):
        self.client.force_login(self.blr_staff)
        # Two rows for the same product (20 in stock): 12 + 12 = 24 > 20, must be rejected as a whole.
        response = self.client.post(reverse('dashboard:stock_ledger:sale_add'), {
            'sale_date': '2026-03-01', 'customer': '', 'notes': '',
            'item_product': [str(self.product.id), str(self.product.id)],
            'item_quantity': ['12', '12'],
            'item_selling_price': ['9.00', '9.00'],
        })
        self.assertEqual(response.status_code, 200)
        self.assertEqual(StockSale.objects.count(), 0)
        stock = BranchStock.objects.get(product=self.product, branch=self.blr)
        self.assertEqual(stock.quantity, 20)  # untouched

    def test_incomplete_item_row_is_silently_skipped(self):
        self.client.force_login(self.blr_staff)
        response = self.client.post(reverse('dashboard:stock_ledger:sale_add'), {
            'sale_date': '2026-03-01', 'customer': '', 'notes': '',
            'item_product': [str(self.product.id), ''],
            'item_quantity': ['2', ''],
            'item_selling_price': ['9.00', ''],
        })
        self.assertEqual(response.status_code, 302)
        self.assertEqual(StockSale.objects.count(), 1)


class OverviewZeroStockTests(StockLedgerTestBase):
    def setUp(self):
        super().setUp()
        OpeningStock.objects.create(product=self.product, branch=self.blr, quantity=10, price=5, created_by=self.admin)
        recompute_branch_stock(self.product.id, self.blr.id)
        StockSale.objects.create(product=self.product, branch=self.blr, quantity=10, price=5, selling_price=8, created_by=self.admin)
        recompute_branch_stock(self.product.id, self.blr.id)

    def test_zero_stock_row_hidden_by_default_and_shown_with_toggle(self):
        self.client.force_login(self.admin)
        response = self.client.get(reverse('dashboard:stock_ledger:overview'))
        self.assertNotIn(self.product.name.encode(), response.content)

        response = self.client.get(reverse('dashboard:stock_ledger:overview'), {'show_zero': '1'})
        self.assertIn(self.product.name.encode(), response.content)

    def test_branch_stock_delete_blocked_when_quantity_nonzero(self):
        BranchStock.objects.create(product=self.product, branch=self.maa, quantity=3)
        row = BranchStock.objects.get(product=self.product, branch=self.maa)
        self.client.force_login(self.admin)
        response = self.client.post(reverse('dashboard:stock_ledger:branch_stock_delete', args=[row.pk]))
        self.assertFalse(response.json()['success'])
        self.assertTrue(BranchStock.objects.filter(pk=row.pk).exists())

    def test_branch_stock_delete_succeeds_at_zero(self):
        row = BranchStock.objects.get(product=self.product, branch=self.blr)
        self.assertEqual(row.quantity, 0)
        self.client.force_login(self.admin)
        response = self.client.post(reverse('dashboard:stock_ledger:branch_stock_delete', args=[row.pk]))
        self.assertTrue(response.json()['success'])
        self.assertFalse(BranchStock.objects.filter(pk=row.pk).exists())


class ProductDeleteTests(StockLedgerTestBase):
    def test_delete_blocked_when_stock_remains(self):
        OpeningStock.objects.create(product=self.product, branch=self.blr, quantity=12, price=5, created_by=self.admin)
        recompute_branch_stock(self.product.id, self.blr.id)

        self.client.force_login(self.admin)
        response = self.client.post(reverse('dashboard:product_delete', args=[self.product.pk]))
        body = response.json()
        self.assertFalse(body['success'])
        self.assertIn('12', body['message'])
        self.assertTrue(Product.objects.filter(pk=self.product.pk).exists())

    def test_delete_blocked_when_transfer_in_progress(self):
        OpeningStock.objects.create(product=self.product, branch=self.blr, quantity=10, price=5, created_by=self.admin)
        recompute_branch_stock(self.product.id, self.blr.id)
        StockTransfer.objects.create(
            product=self.product, from_branch=self.blr, to_branch=self.maa, quantity=5, created_by=self.admin,
        )

        self.client.force_login(self.admin)
        response = self.client.post(reverse('dashboard:product_delete', args=[self.product.pk]))
        self.assertFalse(response.json()['success'])
        self.assertTrue(Product.objects.filter(pk=self.product.pk).exists())

    def test_delete_succeeds_at_zero_stock_and_preserves_history_via_snapshot(self):
        purchase = StockPurchase.objects.create(
            product=self.product, branch=self.blr, quantity=10, price=5, created_by=self.admin,
        )
        recompute_branch_stock(self.product.id, self.blr.id)
        StockSale.objects.create(
            product=self.product, branch=self.blr, quantity=10, price=5, selling_price=9, created_by=self.admin,
        )
        recompute_branch_stock(self.product.id, self.blr.id)

        stock = BranchStock.objects.get(product=self.product, branch=self.blr)
        self.assertEqual(stock.quantity, 0)
        original_name = self.product.name

        self.client.force_login(self.admin)
        response = self.client.post(reverse('dashboard:product_delete', args=[self.product.pk]))
        self.assertTrue(response.json()['success'])
        self.assertFalse(Product.objects.filter(pk=self.product.pk).exists())
        self.assertFalse(BranchStock.objects.filter(product_id=self.product.pk).exists())

        purchase.refresh_from_db()
        self.assertIsNone(purchase.product_id)
        self.assertEqual(purchase.product_name_snapshot, original_name)


class StockDashboardPermissionTests(StockLedgerTestBase):
    """Overview is Admin-only; the new company-wide Dashboard is for everybody."""

    def test_overview_is_admin_only(self):
        for user in (self.blr_staff, self.viewer):
            self.client.force_login(user)
            response = self.client.get(reverse('dashboard:stock_ledger:overview'))
            self.assertEqual(response.status_code, 403)

        self.client.force_login(self.admin)
        response = self.client.get(reverse('dashboard:stock_ledger:overview'))
        self.assertEqual(response.status_code, 200)

    def test_dashboard_is_visible_to_every_role(self):
        for user in (self.admin, self.blr_staff, self.viewer):
            self.client.force_login(user)
            response = self.client.get(reverse('dashboard:stock_ledger:dashboard'))
            self.assertEqual(response.status_code, 200)

    def test_non_admin_login_redirects_to_dashboard_not_overview(self):
        response = self.client.post(reverse('dashboard:login'), {
            'username': 'blrstaff', 'password': 'testpass123',
        })
        self.assertRedirects(response, reverse('dashboard:stock_ledger:dashboard'))


class StockDashboardMatrixTests(StockLedgerTestBase):
    """historical_stock_matrix reconstructs stock as of a given date from
    ledger entries dated on or before it, independent of BranchStock's
    always-current live cache."""

    def test_matrix_reflects_only_entries_on_or_before_as_of_date(self):
        OpeningStock.objects.create(
            product=self.product, branch=self.blr, quantity=100, price=10,
            effective_date=date(2026, 1, 1), created_by=self.admin,
        )
        StockPurchase.objects.create(
            product=self.product, branch=self.blr, quantity=20, price=10,
            purchase_date=date(2026, 2, 1), created_by=self.admin,
        )
        StockSale.objects.create(
            product=self.product, branch=self.blr, quantity=15, price=10, selling_price=15,
            sale_date=date(2026, 3, 1), created_by=self.admin,
        )

        # Before the purchase: only the opening balance counts.
        matrix = historical_stock_matrix(date(2026, 1, 15))
        self.assertEqual(matrix.get((self.product.id, self.blr.id), 0), 100)

        # After the purchase, before the sale.
        matrix = historical_stock_matrix(date(2026, 2, 15))
        self.assertEqual(matrix.get((self.product.id, self.blr.id), 0), 120)

        # After all three entries.
        matrix = historical_stock_matrix(date(2026, 3, 15))
        self.assertEqual(matrix.get((self.product.id, self.blr.id), 0), 105)

        # Before anything happened.
        matrix = historical_stock_matrix(date(2025, 12, 1))
        self.assertEqual(matrix.get((self.product.id, self.blr.id), 0), 0)

    def test_transfer_in_transit_counts_at_neither_branch(self):
        OpeningStock.objects.create(
            product=self.product, branch=self.blr, quantity=50, price=10,
            effective_date=date(2026, 1, 1), created_by=self.admin,
        )
        recompute_branch_stock(self.product.id, self.blr.id)
        transfer = StockTransfer.objects.create(
            product=self.product, from_branch=self.blr, to_branch=self.maa, quantity=20, created_by=self.admin,
        )
        transfer.status = StockTransfer.DISPATCHED
        transfer.dispatched_at = timezone.make_aware(datetime(2026, 2, 1, 10, 0, 0))
        transfer.save()

        # Between dispatch and receipt: gone from BLR, not yet arrived at MAA.
        matrix = historical_stock_matrix(date(2026, 2, 10))
        self.assertEqual(matrix.get((self.product.id, self.blr.id), 0), 30)
        self.assertEqual(matrix.get((self.product.id, self.maa.id), 0), 0)

        transfer.status = StockTransfer.RECEIVED
        transfer.received_quantity = 20
        transfer.received_at = timezone.make_aware(datetime(2026, 2, 5, 10, 0, 0))
        transfer.save()

        matrix = historical_stock_matrix(date(2026, 2, 10))
        self.assertEqual(matrix.get((self.product.id, self.blr.id), 0), 30)
        self.assertEqual(matrix.get((self.product.id, self.maa.id), 0), 20)

    def test_dashboard_view_renders_pivoted_columns(self):
        OpeningStock.objects.create(
            product=self.product, branch=self.blr, quantity=40, price=10,
            effective_date=date(2026, 1, 1), created_by=self.admin,
        )
        self.client.force_login(self.admin)
        response = self.client.get(reverse('dashboard:stock_ledger:dashboard'), {'date': '2026-01-15'})
        self.assertEqual(response.status_code, 200)
        content = response.content.decode()
        self.assertIn(self.product.name, content)
        self.assertIn(self.blr.name, content)
        self.assertIn(self.maa.name, content)
