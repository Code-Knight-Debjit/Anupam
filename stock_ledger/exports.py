from io import BytesIO

from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font
from reportlab.lib import colors
from reportlab.lib.pagesizes import landscape, A4
from reportlab.lib.units import cm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet


def _xlsx_response(filename, headers, rows):
    wb = Workbook()
    ws = wb.active
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)
    for row in rows:
        ws.append(row)
    for i, _ in enumerate(headers, start=1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = 18

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    response = HttpResponse(
        buf.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


def _pdf_response(filename, title, headers, rows):
    buf = BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=landscape(A4), topMargin=1.5 * cm, bottomMargin=1.5 * cm)
    styles = getSampleStyleSheet()
    elements = [Paragraph(title, styles['Title']), Spacer(1, 0.5 * cm)]

    table_data = [headers] + [[str(cell) for cell in row] for row in rows]
    table = Table(table_data, repeatRows=1)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#0e0e11')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTSIZE', (0, 0), (-1, -1), 8),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#dddddd')),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f6f5f7')]),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
    ]))
    elements.append(table)
    doc.build(elements)
    buf.seek(0)
    response = HttpResponse(buf.read(), content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


def export_branch_stock(rows, include_money, fmt):
    """rows: iterable of BranchStock with .product/.branch selected, plus an
    injected `unit_price`/`value` pair from the caller when include_money."""
    headers = ['Product', 'SKU', 'Brand', 'Branch', 'Quantity (Pcs)']
    if include_money:
        headers += ['Unit Price', 'Total Value']

    data = []
    for r in rows:
        line = [r.product.name, r.product.sku, r.product.brand.name if r.product.brand_id else '', r.branch.name, r.quantity]
        if include_money:
            line += [f'{r.unit_price:.2f}', f'{r.value:.2f}']
        data.append(line)

    if fmt == 'pdf':
        return _pdf_response('stock_overview.pdf', 'Stock Ledger — Current Stock', headers, data)
    return _xlsx_response('stock_overview.xlsx', headers, data)


def export_ledger_history(kind, rows, include_money, fmt):
    """kind: 'opening' | 'purchase' | 'sale'"""
    if kind == 'opening':
        headers = ['Product', 'Branch', 'Quantity', 'Effective Date']
        if include_money:
            headers += ['Price', 'Total Price']
        data = []
        for r in rows:
            line = [r.product.name, r.branch.name, r.quantity, r.effective_date]
            if include_money:
                line += [f'{r.price:.2f}', f'{r.total_price:.2f}']
            data.append(line)
    elif kind == 'purchase':
        headers = ['Product', 'Branch', 'Quantity', 'Purchase Date', 'Supplier']
        if include_money:
            headers += ['Price', 'Total Price']
        data = []
        for r in rows:
            line = [r.product.name, r.branch.name, r.quantity, r.purchase_date, r.supplier]
            if include_money:
                line += [f'{r.price:.2f}', f'{r.total_price:.2f}']
            data.append(line)
    else:  # sale
        headers = ['Product', 'Branch', 'Quantity', 'Sale Date', 'Customer']
        if include_money:
            headers += ['Cost Price', 'Selling Price', 'Profit']
        data = []
        for r in rows:
            line = [r.product.name, r.branch.name, r.quantity, r.sale_date, r.customer]
            if include_money:
                line += [f'{r.price:.2f}', f'{r.selling_price:.2f}', f'{r.total_profit:.2f}']
            data.append(line)

    title = f'Stock Ledger — {kind.title()} History'
    filename_base = f'{kind}_history'
    if fmt == 'pdf':
        return _pdf_response(f'{filename_base}.pdf', title, headers, data)
    return _xlsx_response(f'{filename_base}.xlsx', headers, data)
